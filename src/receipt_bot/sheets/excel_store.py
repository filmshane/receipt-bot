from __future__ import annotations

from contextlib import contextmanager
from pathlib import Path
from typing import Any, Dict, Iterator, List, Optional

import portalocker
from openpyxl import Workbook, load_workbook

from receipt_bot.models import ExpenseRow
from receipt_bot.sheets.headers import HEADERS


class ExcelExpenseStore:
    """Cross-platform Excel store (Windows + Linux/Ubuntu)."""

    def __init__(self, path: str | Path, sheet_name: str = "Expenses"):
        self.path = Path(path).expanduser().resolve()
        self.sheet_name = sheet_name

    def ensure_workbook(self) -> None:
        self.path.parent.mkdir(parents=True, exist_ok=True)
        if self.path.exists():
            return
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = self.sheet_name
        ws.append(list(HEADERS))
        wb.save(self.path)

    def _lock_path(self) -> Path:
        return self.path.parent / (self.path.name + ".lock")

    @contextmanager
    def _with_lock(self) -> Iterator[None]:
        """Exclusive lock that works on Ubuntu (fcntl) and Windows."""
        self.ensure_workbook()
        lock_path = self._lock_path()
        lock_path.parent.mkdir(parents=True, exist_ok=True)
        if not lock_path.exists():
            lock_path.write_text("", encoding="utf-8")
        # binary mode + LOCK_EX is portable; avoid timeout kw (no-op/warn on Win)
        with open(lock_path, "a+b") as fh:
            portalocker.lock(fh, portalocker.LOCK_EX)
            try:
                yield
            finally:
                portalocker.unlock(fh)

    def append_expense(self, row: ExpenseRow) -> None:
        with self._with_lock():
            if self.find_by_message_id_unlocked(row.message_id):
                return
            wb = load_workbook(self.path)
            ws = wb[self.sheet_name] if self.sheet_name in wb.sheetnames else wb.active
            assert ws is not None
            if ws.max_row == 0 or (ws.max_row >= 1 and ws["A1"].value is None):
                ws.append(list(HEADERS))
            first = [c.value for c in ws[1]]
            if first[: len(HEADERS)] != list(HEADERS):
                if all(x is None for x in first):
                    for i, h in enumerate(HEADERS, start=1):
                        ws.cell(1, i, h)
            ws.append(row.to_excel_row())
            wb.save(self.path)

    def find_by_message_id(self, message_id: int) -> Optional[Dict[str, Any]]:
        with self._with_lock():
            return self.find_by_message_id_unlocked(message_id)

    def find_by_message_id_unlocked(self, message_id: int) -> Optional[Dict[str, Any]]:
        self.ensure_workbook()
        wb = load_workbook(self.path, read_only=True, data_only=True)
        try:
            ws = wb[self.sheet_name] if self.sheet_name in wb.sheetnames else wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return None
            header = [str(h) if h is not None else "" for h in rows[0]]
            try:
                idx = header.index("MessageId")
            except ValueError:
                idx = list(HEADERS).index("MessageId")
            for r in rows[1:]:
                if r is None:
                    continue
                vals = list(r)
                if idx < len(vals) and vals[idx] is not None:
                    try:
                        if int(vals[idx]) == int(message_id):
                            d: Dict[str, Any] = {}
                            for i, h in enumerate(header):
                                if h:
                                    d[h] = vals[i] if i < len(vals) else ""
                            return d
                    except (TypeError, ValueError):
                        continue
            return None
        finally:
            wb.close()

    def _all_dicts_unlocked(self) -> List[Dict[str, Any]]:
        self.ensure_workbook()
        wb = load_workbook(self.path, read_only=True, data_only=True)
        try:
            ws = wb[self.sheet_name] if self.sheet_name in wb.sheetnames else wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return []
            header = [str(h) if h is not None else "" for h in rows[0]]
            out: List[Dict[str, Any]] = []
            for r in rows[1:]:
                if not r or all(c is None or c == "" for c in r):
                    continue
                d: Dict[str, Any] = {}
                for i, h in enumerate(header):
                    if h:
                        d[h] = r[i] if i < len(r) else ""
                out.append(d)
            return out
        finally:
            wb.close()

    def query(
        self,
        *,
        category: Optional[str] = None,
        user_id: Optional[int] = None,
        date_from: Optional[str] = None,
        date_to: Optional[str] = None,
    ) -> List[Dict[str, Any]]:
        with self._with_lock():
            rows = self._all_dicts_unlocked()
        result = []
        for d in rows:
            if category and str(d.get("Category", "")).lower() != category.lower():
                continue
            if user_id is not None:
                try:
                    if int(d.get("TelegramUserId") or 0) != int(user_id):
                        continue
                except (TypeError, ValueError):
                    continue
            ed = str(d.get("ExpenseDate") or "")
            if date_from and ed < date_from:
                continue
            if date_to and ed > date_to:
                continue
            result.append(d)
        return result

    def sum_total(self, **filters: Any) -> float:
        rows = self.query(**filters)
        total = 0.0
        for d in rows:
            try:
                total += float(d.get("Total") or 0)
            except (TypeError, ValueError):
                continue
        return round(total, 2)

    def list_recent(self, n: int = 10, user_id: Optional[int] = None) -> List[Dict[str, Any]]:
        rows = self.query(user_id=user_id)
        return rows[-n:]

    def update_cfo_sent(self, message_id: int, sent_at: str) -> None:
        with self._with_lock():
            wb = load_workbook(self.path)
            ws = wb[self.sheet_name] if self.sheet_name in wb.sheetnames else wb.active
            assert ws is not None
            header = [c.value for c in ws[1]]
            try:
                mid_i = header.index("MessageId") + 1
                cfo_i = header.index("CFOEmailSentAt") + 1
            except ValueError:
                wb.close()
                return
            for row_idx in range(2, ws.max_row + 1):
                val = ws.cell(row_idx, mid_i).value
                if val is not None:
                    try:
                        if int(val) == int(message_id):
                            ws.cell(row_idx, cfo_i, sent_at)
                            break
                    except (TypeError, ValueError):
                        continue
            wb.save(self.path)
